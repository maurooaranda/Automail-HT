# htmails_ss.py: Defines the format and operations of a HT-Mails spreadsheet
# -*- coding: utf-8 -*-

# Copyright (C) 2018-2019. Mauro Aranda

# This file is part of Automail-HT.

# Automail-HT is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.

# Automail-HT is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.

# You should have received a copy of the GNU General Public License
# along with Automail-HT.  If not, see <https://www.gnu.org/licenses/>.


# Imports for handling spreadsheets.
from openpyxl import load_workbook
import spreadsheet

class htmails_ss:    
    def __init__ (self, ss_path):
        try:
            # TODO: Reject spreadsheet if it doesn't adjust to protocol!!
            self.wb = load_workbook (filename = ss_path, read_only = True)
        except:
            print "No pudo abrirse el archivo indicado"
        else:
            self.preferences = {}
            self.preferences["Seconds_wait"] = self.get_seconds_wait ()
            self.preferences["Keywords"] = self.get_mail_keywords ()
            self.preferences["Headers"] = spreadsheet.get_headers (self.wb,
                                                                   "HT-Mails")
            self.preferences["BlackList"] = self.get_blacklist ()

    def get_mail_keywords (self):
        """Get the keywords that are supported."""
        
        ws = self.wb["Instrucciones"]
        r = 2
        
        keywords = []

        next_keyword = ws.cell (row = r, column = 2).value
        while (next_keyword != None):
            keywords.append (next_keyword)
            r += 1
            next_keyword = ws.cell (row = r, column = 2).value

        return keywords
    
    def get_seconds_wait (self):
        """Get the seconds to wait, and default to 10."""
        
        ws = self.wb["Instrucciones"]

        try:
            return int (ws["J2"].value)
        except ValueError:
            return 10

    def get_total_messages (self):
        """Get how many mail templates are needed."""
        
        messages_number = \
            spreadsheet.get_values_until_blankcell (self.wb, "HT-Mails",
                                                    self.preferences["Headers"].index ("Mensaje") + 1,
                                                    2)
        return max (messages_number)

    def get_thread_id (self):
        """Find the thread id if given, or return 'None'."""
        
        ws = self.wb["Instrucciones"]
        
        return str (ws["J3"].value)

    def get_fields (self):
        """Get the fields in HT-Mails.  These contain the data that is needed
        to send the HT-Mail."""
        
        return spreadsheet.get_values_until_blankrow (self.wb, "HT-Mails",
                                                      len (self.preferences["Headers"]), 2)

    def get_blacklist (self):
        """Get users in blacklist."""
        
        return spreadsheet.get_values_until_blankcell (self.wb, "Equipos Omitidos", 1, 2)
