# -*- coding: utf-8 -*-

# Copyright (C) 2018-2019. Mauro Aranda

# This file is part of Automail-HT.

# Automail-HT is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.

# Automail-HT is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.

# You should have received a copy of the GNU General Public License
# along with Automail-HT.  If not, see <https://www.gnu.org/licenses/>.


# spreadsheet.py: Handles operations of spreadsheets, using openpyxl

# Imports for handling *.xls
import openpyxl

def get_headers(workbook, sheet):
    """Gets headers of a SHEET.  Assumes that when a blank cell is found,
    there are no more headers."""
    
    ws = workbook[sheet]

    col = 1

    headers = []

    next_header = ws.cell(row = 1, column = col).value
    while (next_header != None):
        headers.append(next_header)
        col += 1
        next_header = ws.cell(row = 1, column = col).value
        
    return headers

def get_values_until_blankcell(workbook, sheet, col, row_start):
    """Get all values from a SHEET, for a column COL, starting at ROW_START"""
    ws = workbook[sheet]

    r = row_start

    data = []

    next_data = ws.cell(row = r, column = col).value

    while (next_data != None):
        data.append(next_data)
        r += 1
        next_data = ws.cell(row = r, column = col).value

    return data
    
def get_values_until_blankrow(workbook, sheet, cols, row_start):
    """Get all values in rows, until a row is blank"""
    
    ws = workbook[sheet]

    r = row_start
    ret = ()
    
    data = []
    field = []
    i = 0

    data.append([])
    for col in range(cols):
        data[i].append(ws.cell(row = r, column = col + 1).value)

    while (not(all(v is None for v in data[i]))):
        data[i] = tuple(data[i])
        i += 1
        r += 1
        data.append([])
        for col in range(cols):
            data[i].append(ws.cell(row = r, column = col + 1).value)

    # A None tuple was added, so delete it.
    del data[-1]

    return data
